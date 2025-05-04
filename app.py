import streamlit as st
import pandas as pd
from datetime import date
import os
import openpyxl
import shutil
from datetime import datetime
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import numpy as np

# Set page title and configuration
st.set_page_config(page_title="Student Payment Tracker", layout="wide")
st.title("Student Payment Tracker")

# Define file paths
template_file = "resources/Student Payment Tracker.xlsx"
output_dir = "output"

# Create output directory if it doesn't exist
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Function to write data to Excel while preserving template formatting
def write_excel_file(file_path, dataframe):
    # Define the header row - start at row 15 as requested
    HEADER_ROW = 15
    
    # Check if template exists to preserve its formatting
    if os.path.exists(template_file):
        # Make a copy of the template to preserve logos and formatting
        shutil.copy(template_file, file_path)
        # Open the copied template
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        # If template doesn't exist, create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
    
    # ALWAYS write the column headers starting at row 15
    # Set up headers with each column taking two columns
    # Student Name: A15:B15, Payment Amount: C15:D15, Payment Date: E15:F15
    try:
        # Attempt to unmerge first in case they are already merged
        try:
            ws.unmerge_cells(f'A{HEADER_ROW}:B{HEADER_ROW}')
        except:
            pass
        ws.merge_cells(f'A{HEADER_ROW}:B{HEADER_ROW}')
        header_cell = ws.cell(row=HEADER_ROW, column=1)
        header_cell.value = "Student Name"
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.font = Font(bold=True)
        
        try:
            ws.unmerge_cells(f'C{HEADER_ROW}:D{HEADER_ROW}')
        except:
            pass
        ws.merge_cells(f'C{HEADER_ROW}:D{HEADER_ROW}')
        header_cell = ws.cell(row=HEADER_ROW, column=3)
        header_cell.value = "Payment Amount"
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.font = Font(bold=True)
        
        try:
            ws.unmerge_cells(f'E{HEADER_ROW}:F{HEADER_ROW}')
        except:
            pass
        ws.merge_cells(f'E{HEADER_ROW}:F{HEADER_ROW}')
        header_cell = ws.cell(row=HEADER_ROW, column=5)
        header_cell.value = "Payment Date"
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.font = Font(bold=True)
    except Exception as e:
        st.warning(f"Could not set headers: {str(e)}")
    
    # Define cell styling for centering
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Define color fills for alternating rows
    # Range of light colors for better readability
    colors = [
        "FFD6E0", "FFE8D6", "D6EAD6", "D6E0EA", "E0D6EA", 
        "EAD6E0", "EAE0D6", "E0EAD6", "D6EAEA", "E0D6EA"
    ]
    
    # Set column widths if not using template
    if not os.path.exists(template_file):
        for col in "ABCDEF":
            ws.column_dimensions[col].width = 15
    
    # Start filling data from one row after the header
    start_row = HEADER_ROW + 1
    
    # Calculate total payment amount
    total_payment = 0.0
    
    # Fill in data, making each value occupy two columns - ENSURE CORRECT COLUMN ORDER
    for row_idx, row_data in enumerate(dataframe.values, start=start_row):
        # Set row color
        color_idx = (row_idx - start_row) % len(colors)
        row_fill = PatternFill(start_color=colors[color_idx], end_color=colors[color_idx], fill_type="solid")
        
        try:
            # Student Name in A and B (ensure title case) - INDEX 0
            # Always merge these cells regardless of existing content (fix for merged cells)
            try:
                ws.merge_cells(f'A{row_idx}:B{row_idx}')
            except:
                # If cells are already merged, unmerge and then merge again
                try:
                    ws.unmerge_cells(f'A{row_idx}:B{row_idx}')
                    ws.merge_cells(f'A{row_idx}:B{row_idx}')
                except:
                    pass  # If this fails too, just proceed
                
            cell = ws.cell(row=row_idx, column=1)
            student_name_val = str(row_data[0]).title() if row_data[0] is not None else ""
            cell.value = student_name_val
            cell.alignment = center_alignment
            cell.fill = row_fill
        
            # Payment Amount in C and D - add AED prefix - INDEX 1
            try:
                ws.merge_cells(f'C{row_idx}:D{row_idx}')
            except:
                try:
                    ws.unmerge_cells(f'C{row_idx}:D{row_idx}')
                    ws.merge_cells(f'C{row_idx}:D{row_idx}')
                except:
                    pass
                
            cell = ws.cell(row=row_idx, column=3)
            # Handle the payment amount appropriately
            amount_value = row_data[1]  # Payment Amount
            
            # Try to convert to float for accumulating the total
            try:
                if amount_value is not None and str(amount_value).strip() != "":
                    numeric_amount = float(amount_value)
                    total_payment += numeric_amount
                    cell.value = f"AED {numeric_amount:.2f}"
                else:
                    cell.value = "AED 0.00"
            except (ValueError, TypeError):
                cell.value = f"AED {amount_value}"
            
            cell.alignment = center_alignment
            cell.fill = row_fill
            
            # Payment Date in E and F - INDEX 2
            try:
                ws.merge_cells(f'E{row_idx}:F{row_idx}')
            except:
                try:
                    ws.unmerge_cells(f'E{row_idx}:F{row_idx}')
                    ws.merge_cells(f'E{row_idx}:F{row_idx}')
                except:
                    pass
                
            cell = ws.cell(row=row_idx, column=5)
            date_val = str(row_data[2]) if row_data[2] is not None else ""
            cell.value = date_val
            cell.alignment = center_alignment
            cell.fill = row_fill
        except Exception as e:
            st.error(f"Error writing row {row_idx}: {str(e)}")
            continue  # Skip this row if there's an error
    
    # Add a total row ONLY after the last student
    total_row = len(dataframe.values) + start_row  # +start_row to account for headers
    
    # "Total" label in columns A and B
    try:
        ws.merge_cells(f'A{total_row}:B{total_row}')
    except:
        try:
            ws.unmerge_cells(f'A{total_row}:B{total_row}')
            ws.merge_cells(f'A{total_row}:B{total_row}')
        except:
            pass
            
    cell = ws.cell(row=total_row, column=1)
    cell.value = "Total"
    cell.alignment = center_alignment
    cell.font = Font(bold=True)
    
    # Total amount in columns C and D
    try:
        ws.merge_cells(f'C{total_row}:D{total_row}')
    except:
        try:
            ws.unmerge_cells(f'C{total_row}:D{total_row}')
            ws.merge_cells(f'C{total_row}:D{total_row}')
        except:
            pass
            
    cell = ws.cell(row=total_row, column=3)
    cell.value = f"AED {total_payment:.2f}"
    cell.alignment = center_alignment
    cell.font = Font(bold=True)
    
    # Empty cell in columns E and F
    try:
        ws.merge_cells(f'E{total_row}:F{total_row}')
    except:
        try:
            ws.unmerge_cells(f'E{total_row}:F{total_row}')
            ws.merge_cells(f'E{total_row}:F{total_row}')
        except:
            pass
    
    # Save the workbook
    try:
        wb.save(file_path)
    except Exception as e:
        st.error(f"Error saving workbook: {str(e)}")

# Initialize session state to store student data across reruns
if 'student_data' not in st.session_state:
    st.session_state.student_data = pd.DataFrame(columns=["Student Name", "Payment Amount", "Payment Date"])

# Generate a unique filename for this session
if 'current_file' not in st.session_state:
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"{output_dir}/Student_Payment_Tracker_{current_time}.xlsx"
    st.session_state.current_file = excel_file
else:
    excel_file = st.session_state.current_file

# Check if we should try to read from existing template
if not st.session_state.student_data.empty:
    # We already have data, use what's in session state
    df = st.session_state.student_data
else:
    # Try to read from template or start fresh
    if os.path.exists(template_file):
        try:
            # Create default empty DataFrame
            df = pd.DataFrame(columns=["Student Name", "Payment Amount", "Payment Date"])
            
            # Try to read data from existing template file
            try:
                # Read raw data
                raw_df = pd.read_excel(template_file)
                
                # If there's data, extract what we can
                if not raw_df.empty:
                    # Look for data that appears to be student records - skip first 15 rows
                    for idx, row in raw_df.iterrows():
                        # Skip header row(s), rows with "Total", and all rows before row 15
                        if idx < 15 or (isinstance(row.iloc[0], str) and "total" in str(row.iloc[0]).lower()):
                            continue
                            
                        # Extract data from row - handle varying column counts
                        name = str(row.iloc[0]) if len(row) > 0 else ""
                        
                        # Try to get payment amount, which might be in column 1 or 2
                        payment = None
                        if len(row) > 1:
                            try:
                                # Try to convert to float
                                payment = float(str(row.iloc[1]).replace("AED", "").strip())
                            except:
                                # If it fails, try the next column
                                if len(row) > 2:
                                    try:
                                        payment = float(str(row.iloc[2]).replace("AED", "").strip())
                                    except:
                                        payment = 0.0
                                else:
                                    payment = 0.0
                        else:
                            payment = 0.0
                        
                        # Get date from remaining column
                        date_val = ""
                        if len(row) > 2:
                            date_val = str(row.iloc[2])
                        
                        # Add to our dataframe if it looks like a valid entry
                        if name and name.strip() and name.lower() != "total":
                            df = pd.concat([df, pd.DataFrame([{
                                "Student Name": name,
                                "Payment Amount": payment,
                                "Payment Date": date_val
                            }])], ignore_index=True)
            except Exception as e:
                st.warning(f"Could not read existing data from template: {str(e)}")
                # Continue with empty dataframe
                
            # Save the initial data to session state
            st.session_state.student_data = df
        except Exception as e:
            st.error(f"Error initializing data: {str(e)}")
            df = pd.DataFrame(columns=["Student Name", "Payment Amount", "Payment Date"])
            st.session_state.student_data = df
    else:
        # No template, start with empty DataFrame
        df = pd.DataFrame(columns=["Student Name", "Payment Amount", "Payment Date"])
        st.session_state.student_data = df

# Display current file being used
st.info(f"Working with file: {os.path.basename(excel_file)}")

# Create form for data entry
with st.form("payment_form", clear_on_submit=True):
    st.subheader("Enter Payment Information")
    
    # Create two columns for the form layout
    col1, col2 = st.columns(2)
    
    # Input fields in columns
    with col1:
        student_name = st.text_input("Student Name")
    
    with col2:
        payment_amount = st.number_input("Payment Amount (AED)", min_value=0.0, step=0.01, format="%.2f")
    
    # Date picker in its own row spanning two columns
    payment_date = st.date_input("Payment Date", value=date.today())
    
    # Submit button
    submitted = st.form_submit_button("Submit Payment")
    
    if submitted:
        if not student_name:  # Validate student name is not empty
            st.error("Please enter a student name")
        else:
            # Format student name in title case (capitalize first letter of each word)
            formatted_name = student_name.title()
            
            # Format date as YYYY-MM-DD string
            formatted_date = payment_date.strftime('%Y-%m-%d')
            
            # Add new data to dataframe and save to session state
            new_data = {
                "Student Name": formatted_name,
                "Payment Amount": payment_amount,
                "Payment Date": formatted_date
            }
            
            # Append to the session state data
            st.session_state.student_data = pd.concat([
                st.session_state.student_data, 
                pd.DataFrame([new_data])
            ], ignore_index=True)
            
            # Update the local variable too
            df = st.session_state.student_data
            
            # Write the Excel file with all entries
            write_excel_file(excel_file, df)
            
            st.success(f"Payment of AED {payment_amount:.2f} for {formatted_name} has been recorded!")

# Display existing data
st.subheader("Payment Records")
if not df.empty:
    # Try to convert Payment Amount column to numeric, coercing errors to NaN
    display_df = df.copy()
    display_df["Payment Amount"] = pd.to_numeric(display_df["Payment Amount"], errors="coerce")
    
    # Calculate total amount for display with error handling
    try:
        total_amount = display_df["Payment Amount"].sum()
        if pd.isna(total_amount):
            total_amount_str = "AED 0.00 (some values couldn't be summed)"
        else:
            total_amount_str = f"AED {total_amount:.2f}"
    except Exception:
        total_amount_str = "AED 0.00 (error calculating sum)"
    
    # Format payment amount to show AED in the display table
    # Also ensure student names are in title case
    
    # Format student names in title case
    display_df["Student Name"] = display_df["Student Name"].apply(lambda x: str(x).title() if x is not None else "")
    
    # Safely format the payment amount with error handling
    def format_amount(x):
        try:
            if pd.isna(x):
                return "AED 0.00"
            elif isinstance(x, (int, float)):
                return f"AED {x:.2f}"
            else:
                return f"AED {x}" if not str(x).startswith('AED') else x
        except:
            return f"AED {x}" if x is not None else "AED 0.00"
    
    display_df["Payment Amount"] = display_df["Payment Amount"].apply(format_amount)
    
    # Set custom column widths for the dataframe display
    st.dataframe(
        display_df,
        column_config={
            "Student Name": st.column_config.TextColumn("Student Name", width="large"),
            "Payment Amount": st.column_config.TextColumn("Payment Amount", width="medium"),
            "Payment Date": st.column_config.TextColumn("Payment Date", width="medium"),
        },
        use_container_width=True
    )
    
    # Display the total
    st.info(f"Total Amount: {total_amount_str}")
else:
    st.info("No payment records found. Add a payment to get started!")

# Add download button for the Excel file
if os.path.exists(excel_file):
    with open(excel_file, "rb") as file:
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            label="Download Excel File",
            data=file,
            file_name=f"Student_Payment_Tracker_{current_time}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ) 